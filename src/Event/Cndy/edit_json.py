import os
import json
import openpyxl
import sys
import re

# Function to update JSON files with English translations from an Excel workbook
def update_jsons_from_excel(excel_file, json_folder):
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    
    # Process each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        # Extract index number from sheet name (e.g., hl_<index>)
        match = re.match(r"hl_(\d+)", sheet_name)
        if not match:
            print(f"Skipping sheet '{sheet_name}' (invalid format)")
            continue
        
        index_number = match.group(1)
        # Construct pattern to match corresponding JSON file
        pattern = re.compile(rf"^hl_{index_number}_.*\.cndy\.json$")
        
        # Search for the corresponding JSON file in the specified folder
        json_file_path = None
        for filename in os.listdir(json_folder):
            if pattern.match(filename):
                json_file_path = os.path.join(json_folder, filename)
                break
        
        if not json_file_path:
            print(f"No JSON file found for sheet '{sheet_name}'")
            continue

        # Load JSON data
        with open(json_file_path, 'r', encoding='utf-8') as json_file:
            data = json.load(json_file)

        # Check if JSON has a 'Strings' field
        if 'Strings' not in data:
            print(f"No 'Strings' field in {json_file_path}")
            continue

        # Read translations from the sheet and apply them to the JSON
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):  # Skip header row, use first 3 cols
            string_no, jp_text, en_text = row
            if string_no is None:
                continue

            # Convert `string_no` to an integer if it’s a float
            try:
                string_no = int(string_no)
            except (ValueError, TypeError):
                print(f"Invalid string number '{string_no}' in sheet '{sheet_name}'")
                continue

            # Ensure `string_no` is within the range of available strings
            if string_no < 0 or string_no >= len(data['Strings']):
                print(f"String number {string_no} out of range in {json_file_path}")
                continue

            # Replace Japanese string with English if English translation exists
            if en_text and en_text.strip():
                data['Strings'][string_no] = en_text
            else:
                data['Strings'][string_no] = jp_text  # Keep original if no translation provided

        # Save the updated JSON data back to file
        with open(json_file_path, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
        
        print(f"Updated {json_file_path}")

# Main script
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py <Excel File> <JSON Folder>")
        sys.exit(1)

    excel_file = sys.argv[1]
    json_folder = sys.argv[2]

    update_jsons_from_excel(excel_file, json_folder)
